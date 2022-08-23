from openpyxl import load_workbook
import xlsxwriter
from pathlib import Path


# 操作Excel的工具类
class Excel():
    def __init__(self, type, file_name):
        # 读取excel
        if type == 'r':
            # 打开文件
            self.workbook = load_workbook(file_name)
            # 工作簿中存在所有sheet表名
            self.sheet_names = self.workbook.sheetnames
            # 装载所有数据的list
            self.list_data = []
            # 将测试数据内调用的方法，改编成自定义里面的变量
            self.dict_data = []
        # 写入excel
        elif type == 'w':
            self.workbook = xlsxwriter.Workbook(file_name)
        # 获取到所有的sheet_names,sheetl,sheet2获取到所有，获取到的是一个list

    def read(self):
        # 循环编辑
        for sheet_name in self.sheet_names:
            # 通过sheetname获取指定的工作表
            sheet = self.workbook[sheet_name]
            # 按行读取数据
            for i in sheet.values:
                # 将每一行的内容添加进去
                self.list_data.append(list(i))
            return self.list_data

    def general_data(self):
        datas = self.read()
        gengeral = []
        for i in datas[1:]:
            dict_data = {}
            for j in range(0, len(datas[0])):
                dict_data[datas[0][j]] = i[j]
            if dict_data['datas'] != None:
                jsondata = {}
                for i in dict_data['datas'].replace("{", "").replace("}", "").split("\n"):
                    re = i.replace("'", "").replace(":", "").replace(",", "")
                    if len(re.split()) > 0 and len(re.split()) < 2:
                        jsondata[re.split()[0]] = ''
                    elif len(re.split()) > 1:
                        jsondata[re.split()[0]] = re.split()[1]
                dict_data['datas'] = jsondata
            gengeral.append(dict_data)
        return gengeral


if __name__ == '__main__':
    # file = '../otherdata/interface_general.xlsx'
    # e = Excel('r', file)
    # # list_read=e.read()
    # e.general_data()
    # # print(list_read)
    file_name = '../otherdata/interface_general.xlsx'
    if Path(file_name).is_file():
        datas = Excel('r', file_name).general_data()
        print(datas)
        # return datas
