import operator
from openpyxl import load_workbook
import xlsxwriter
from pathlib import Path

# 操作Excel的工具类
class Excel():
# 初始化方法 参数type：为r是读取excel，为w是写入excel获取不同的实例，参数file_name是将要读取的文件
    def __init__(self,type,file_name):
        #读取excel
        if type=='r':
            #打开文件
            self.workbook=load_workbook(file_name)
            # 工作簿中存在所有sheet表名
            self.sheet_names=self.workbook.sheetnames
            #装载所有数据的list
            self.list_data=[]
            #将测试数据内调用的方法，改编成自定义里面的变量
            self.dict_data=[]
        #写入excel
        elif type=='w':
            self.workbook=xlsxwriter.Workbook(file_name)
        #获取到所有的sheet_names,sheetl,sheet2获取到所有，获取到的是一个list
    def read(self):
        #循环编辑
        for sheet_name in self.sheet_names:
            #通过sheetname获取指定的工作表
            sheet=self.workbook[sheet_name]
            #按行读取数据
            for i in sheet.values:
                # 将每一行的内容添加进去
                self.list_data.append(list(i))
            return self.list_data
def element_tojson(elemnt):
    elements={}
    # 将元素和接口等信息组成key和value的形式方便进行查询
    for e in elemnt:
        elements[e[0]]={'type':e[1],'url':e[2]}
    return elements
"""
1.讲excel头部替换成英文的
2.处理成json格式
"""
def datatodict(data):
    header={
        '用例编号':'id',
        '用例标题':'title',
        '前置条件':'condition',
        '测试功能点':'testdot',
        '测试步骤':'step',
        '操作':'keyword',
        '页面':'page',
        '元素':'element',
        '测试数据':'data',
        '预期结果':'expected',
        '设计者':'designer',
        '结果':'score',
        '备注':'remark'
    }
    head=[]
    list_dict_data=[]
    for d in data[0]:
        # 获取到英文的头部内容如果为中文，则替换成英文 进行改成一个k
        # 传入两个参数的作用是 查到则返回查到的数据查不到则返回传入的原数据
        d=header.get(d,d)
        # 将去除的头部英文装进list中
        head.append(d)
    # 获取到数据进行切片处理，0坐标为标题，1坐标是头部
    for b in data[1:]:
        # 头部和内容拼接为json串
        dict_data={}
        for i in range(len(head)):
            # 判断类型，如果不进行判断会出现str的错误，strip去除空格也有转str的用法
            if isinstance(b[i],str):
                dict_data[head[i]]=b[i].strip()
            else:
                dict_data[head[i]]=b[i]
        #list里面是字典格式
        list_dict_data.append(dict_data)
    return list_dict_data

#dict格式的数据处理为测试套件格式
def suite_format(data):
    #用例套件list
    testsuite=[]
    #每个用例的testcase
    testcase={}
    #得到用例的所有数据
    # 循环遍历判断里面是不是一组用例生成用例集
    for d in data:
        # 判断用例有没有标题，没有标题则认为是统一用例，有标题则认为是第二条第三条用例依次类推
        if d['id'].strip():
            # 判断是否为空
            if testcase:
                # 不为空则只认为用例直接添加到list里面
                # 将testcase置空
                testcase={}
            # 这里生成了用例的标题行，里面没有step
            for key in ('id','title','condition','testdot','designer','remark'):
                # test[key] 为id等值，d[key]为内容值
                testcase[key]=d[key]
            #添加steps字段，并设置为list
            testcase['steps']=[]
            #取出的是测试步骤第一步第二部
        step={}
        # 步骤里面添加control字段
        step['control']=''
        #步骤
        step['no']=str(d['step'])
        #去除这些对应的内容放入step里面
        for key in ('testdot','keyword','element','data','expected','output','score','remark'):
            # 获取用例内容字段进行拼接
            step[key]=d.get(key,'')
            # 仅作为测试结果输出时，保持原样
            # 进行的操作
        step['_keyword']=d['keyword']
        # 对应的key取值
        step['_element']=d['element']
        #测试数据
        step['_data']=d['data']
        #预期结果
        step['_expected']=d.get('expected','')
        #输出结果
        step['_output']=''
        #测试结果信息
        step['_resultinfo']=''
        #添加测试步骤
        testcase['steps'].append(step)
        testsuite.append(testcase)
        #返回处理好的值
        return  testsuite
# 判断当前目录是否存在
def mkdir(p):
    path=Path(p)
    # 如果文件不存在 则创建
    if not path.is_dir():
        path.mkdir()
#获取接口返回值的格式，在httpcaps.py中进行调用
def compare_key_value(json_p):
    list_key=[]
    def getkey_value_all(input_json={}):
        # isinstance() 函数来判断一个对象是否是一个已知的类型
        if isinstance(input_json,dict):
            # keys() 函数以列表返回一个字典所有的键
            for key in input_json.keys():
                # get() 函数返回指定键的值，如果值不在字典中返回默认值。
                key_value=input_json.get(key)
                if isinstance(key_value,dict):
                    getkey_value_all(key_value)
                elif isinstance(key_value,list):
                    for json_array in key_value:
                        getkey_value_all(json_array)
                else:
                    # 对象类型的key
                    list_key.append(str(key))
                    pass
            # 对象类型的key
            list_key.append(str(key))
        elif isinstance(input_json,list):
            for input_json_arrary in input_json:
                getkey_value_all(input_json_arrary)
    getkey_value_all(json_p)
    return list_key
#写入token
def writetoken(token):
    path=Path('config')/('txt_final.txt')
    # 写入token和普通常量
    f=open(path,'a')
    f.write(token)
    f.close()
    return
# 对比两个json的函数
def iscompare_json(sub,parent):
    #将json内容传入获取key值
    a1=compare_key_value(sub)
    a2=compare_key_value(parent)
    #两个key值进行对比
    flag=operator.eq(a1,a2)
    # 一致则通过
    if flag==True:
        return 'Pass'
    else:
        return 'Fail'
if __name__=='__main__':
    # file='../element/elements.xlsx'
    # e=Excel('r',file)
    # list_read=e.read()
    # ele=element_tojson(list_read)
    #
    # test_case='../testcase/testcase.xlsx'
    # e_case=Excel('r',test_case)
    # re=e_case.read()
    # data=datatodict(re)
    # testsuite=suite_format(data)
    # print(testsuite[0]['steps'][0]['data'])
    sub = {'phone': '17547817934', 'type': '1'}
    parent = {'phone': '17547817934', 'sort': '1'}
    print(iscompare_json(sub, parent))